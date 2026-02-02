import csv
import email
import os
import re
import imaplib
from io import StringIO

import openpyxl
from openpyxl import load_workbook
from datetime import datetime, timedelta
import random
import pandas as pd
from datetime import datetime
import string
from robot.libraries.BuiltIn import BuiltIn
from robot.api.deco import keyword
import json
import time
from pathlib import Path
from dateutil.relativedelta import relativedelta
from typing import List
from decimal import Decimal, ROUND_HALF_UP

def read_test_data(file_path, sheet_name, tc_id):

    workbook = load_workbook(file_path)
    sheet = workbook[sheet_name]

    headers = []

    for cell in sheet[1]:
        headers.append(cell.value)

    for row in sheet.iter_rows(min_row = 2, values_only=True):
        current_tc_id = row[0]

        if current_tc_id == tc_id:
            data = {}

            index = 0
            while index < len(headers):
                data[headers[index]] = row[index]
                index += 1
            return data
    raise Exception(f"Test Case {tc_id} not found inn excel")

@keyword
def fetch_testdata_by_id(file_path, target_id):
    global workbook
    try:
        workbook = load_workbook(file_path)
        sheet = workbook.active
        header = [cell.value for cell in sheet[1]]

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0] == target_id:
                data_dict = {}
                for col_name, value in zip(header, row):
                    if ',' in str(value):
                        data_dict[col_name] = [item.strip() for item in value.split(',')]
                    else:
                        data_dict[col_name] = value
                return data_dict

        # If the target_id is not found, you can raise an exception or return a specific value
        raise ValueError(f"Target ID '{target_id}' not found in the Excel file.")
    except Exception as e:
        print(f"Error reading Excel file: {e}")

def no_of_months_ahead(a):
    current_date = datetime.now()

    # Handle case where 'a' is 'NULL'
    if a == 'NULL':
        raise ValueError("The input value 'a' cannot be 'NULL'")

    try:
        a = int(a)  # Convert input to integer
    except ValueError:
        raise ValueError(
            f"Invalid value for a: {a}. Must be an integer or a valid string representation of an integer.")

    # Calculate new month and year
    total_months = current_date.month + a
    new_year = current_date.year + (total_months - 1) // 12
    new_month = ((total_months - 1) % 12) + 1

    # Adjust for negative months
    if total_months <= 0:
        new_year = current_date.year + (total_months - 1) // 12
        new_month = ((total_months - 1) % 12) + 1
        if new_month == 0:
            new_month = 12
            new_year -= 1

    # List of days in each month (accounting for leap year)
    days_in_month = [
        31,  # January
        29 if new_year % 4 == 0 and (new_year % 100 != 0 or new_year % 400 == 0) else 28,  # February
        31,  # March
        30,  # April
        31,  # May
        30,  # June
        31,  # July
        31,  # August
        30,  # September
        31,  # October
        30,  # November
        31  # December
    ]

    # Determine the correct day
    new_day = min(current_date.day, days_in_month[new_month - 1])

    # Create the new date
    new_date = datetime(new_year, new_month, new_day)

    # Format the output
    print("Current Date:", current_date.strftime("%d-%m-%Y"))
    print(f"New Date after adding {a} months:", new_date.strftime("%Y-%m-%d"))
    print("Current month:", current_date.strftime("%B"))

    # Prepare return values
    monthnew = new_date.strftime("%B")
    daynew = new_date.strftime("%d").lstrip('0')  # Remove leading zero
    Yearnew = new_date.strftime("%Y")

    print("New Date (formatted):", daynew, monthnew, Yearnew)

    mainlist = [daynew, monthnew, Yearnew]
    return mainlist

from datetime import datetime

def display_date_formatter(date_input1, month_offset=0):
    try:
        date_input1 = str(date_input1)

        def parse_date(date_input):
            if isinstance(date_input, str):
                try:
                    return datetime.strptime(date_input, "%d/%m/%Y")
                except ValueError:
                    return datetime.strptime(date_input, "%Y-%m-%d %H:%M:%S")
            return date_input

        dt1 = parse_date(date_input1)

        today = datetime.today()
        base_year = today.year
        base_month = today.month + int(month_offset)

        while base_month > 12:
            base_month -= 12
            base_year += 1

        next_month = base_month + 1
        next_year = base_year
        if next_month > 12:
            next_month = 1
            next_year += 1

        def is_valid_month(dt):
            return (
                (dt.year == base_year and dt.month == base_month) or
                (dt.year == next_year and dt.month == next_month)
            )

        date1_found = is_valid_month(dt1)

        return {
            "status": "OK",
            "display_date": dt1.strftime("%a %b %d %Y"),
            "display_date1": dt1.strftime("%a, "),
            "date_found": date1_found,
            "year": dt1.year,
            "month": dt1.month,
            "day": dt1.day,
            "base_year": base_year,
            "base_month": base_month,
            "dept_date_verified": f"{dt1.strftime('%a,')} {dt1.day} {dt1.strftime('%b')} {dt1.year}"
        }

    except Exception as e:
        return {
            "status": "ERROR",
            "msg": str(e),
            "date_found": False
        }
